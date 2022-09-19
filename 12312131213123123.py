# Написать программу для работы с данными о людях

import os
import sys
import openpyxl
import datetime


class Person:
    def __init__(self, name, surname, patronymic, birthday, data_death):
        self.name = name
        self.surname = surname
        self.patronymic = patronymic
        self.birthday = birthday
        self.data_death = data_death

    def __str__(self):
        return f'Имя: {self.name}, Фамилия: {self.surname}, Отчество: {self.patronymic}, Дата рождения: {self.birthday}, Дата смерти: {self.data_death}'

    def __repr__(self):
        return f'Имя: {self.name}, Фамилия: {self.surname}, Отчество: {self.patronymic}, Дата рождения: {self.birthday}, Дата смерти: {self.data_death}'


class PersonList:
    def __init__(self):
        self.persons = []

    def __str__(self):
        return f'Список людей: {self.persons}'

    def __repr__(self):
        return f'Список людей: {self.persons}'

    def add_person(self, person):
        self.persons.append(person)

    def save(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Люди'
        sheet.append(['Имя', 'Фамилия', 'Отчество', 'Дата рождения', 'Дата смерти'])
        for person in self.persons:
            sheet.append([person.name, person.surname, person.patronymic, person.birthday, person.data_death])
        wb.save('persons.xlsx')

    def load(self):
        wb = openpyxl.load_workbook('persons.xlsx')
        sheet = wb['Люди']
        for row in sheet.iter_rows(min_row=2):
            person = Person(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
            self.persons.append(person)

    def find_person(self, surname):
        for person in self.persons:
            if person.surname == surname:
                return person
        return None

    def delete_person(self, surname):
        person = self.find_person(surname)
        if person is not None:
            self.persons.remove(person)
            return True
        return False