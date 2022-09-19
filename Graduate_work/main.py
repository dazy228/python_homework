# Написать программу для работы с данными о людях

import os
import openpyxl
import datetime


class Person(object):
    def __init__(self, name, surname, patronymic, birthday, data_death):
        self.name = name
        self.surname = surname
        self.patronymic = patronymic
        self.birthday = birthday
        self.data_death = data_death

    def __str__(self):
        return f'Имя: {self.name}, Фамилия: {self.surname}, Отчество: {self.patronymic}, Возраст: {"  "}, Дата рождения: {self.birthday}, Дата смерти: {self.data_death}'.title()

    def get_full_name(self):
        return f'{self.surname} {self.name} {self.patronymic}'.title()


class ValidateDate(Person):
    def valid_date(self):
        birthday = self.birthday
        valid_data_list = []
        birthday = birthday.split('.')
        for i in birthday:
            valid_data_list.append(int(i))
        return valid_data_list


class PersonList(object):
    def __init__(self):
        self.persons = []

    def __str__(self):
        return f'Список людей: {self.persons}'

    def __repr__(self):
        return f'Список людей: {self.persons}'

    def add_person(self, person):
        self.persons.append(person)

    def save(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Люди'
        sheet.append(['Имя', 'Фамилия', 'Отчество', 'Дата рождения', 'Дата смерти'])
        for person in self.persons:
            sheet.append([person.name, person.surname, person.patronymic, person.birthday, person.data_death])
        wb.save('persons.xlsx')

    def load(self):
        wb = openpyxl.load_workbook('persons.xlsx')
        sheet = wb['Люди']
        for row in sheet.iter_rows(min_row=2):
            person = Person(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
            self.persons.append(person)

    def get_info(self):
        for person in self.persons:
            print(person)

    def find_persons(self, search_name):
        persons = []
        for person in self.persons:
            if search_name in person.get_full_name():
                persons.append(person)
        return persons

    def find_person(self, surname):
        for person in self.persons:
            if person.surname == surname:
                return person
        return None

    def delete_person(self, surname):
        person = self.find_person(surname)
        if person is not None:
            self.persons.remove(person)
            return True
        return False


def main():
    person_list = PersonList()
    if os.path.exists('persons.xlsx'):
        person_list.load()
    while True:
        print('-' * 30)
        print('1. Добавить человека')
        print('2. Найти человека')
        print('3. Удалить человека')
        print('4. Сохранить')
        print('5. Вывести базу')
        print('6. Выйти')
        print('-' * 30)
        choice = input('Выберите пункт меню: ')
        if choice == '1':
            while True:
                name = input('Введите имя: ')
                if name == '':
                    print('Имя не может быть пустым')
                else:
                    break
            surname = input('Введите фамилию: ')
            patronymic = input('Введите отчество: ')
            while True:
                birthday = input('Введите дату рождения: ')
                try:
                    birthday = datetime.datetime.strptime(birthday, '%d.%m.%Y').strftime('%d.%m.%Y')
                    break
                except ValueError:
                    print('Неверный формат даты')
            while True:
                data_death = input('Введите дату смерти: ')
                try:
                    data_death = datetime.datetime.strptime(data_death, '%d.%m.%Y').strftime('%d.%m.%Y')
                    break
                except ValueError:
                    print('Неверный формат даты')
            person = Person(name, surname, patronymic, birthday, data_death)
            person_list.add_person(person)
        elif choice == '2':
            search_name = input('Введите слово по которому будет произведён поиск: ')
            person = person_list.find_persons(search_name)
            if person is not None:
                for i in person:
                    print(i)
            else:
                print('Человек не найден')
        elif choice == '3':
            surname = input('Введите фамилию: ')
            if person_list.delete_person(surname):
                print('Человек удален')
            else:
                print('Человек не найден')
        elif choice == '4':
            person_list.save()
        elif choice == '5':
            person_list.get_info()
        elif choice == '6':
            break
        else:
            print('Неверный пункт меню')


if __name__ == '__main__':
    main()
