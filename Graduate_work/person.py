import openpyxl
import datetime


class Person(object):
    # Атрибуты для корректного масштабирования выводимой информации о обьекте
    len_name = 0
    len_surname = 0
    len_patronymic = 0

    def __init__(self, name, surname, patronymic, sex, birthday, data_death):
        self.name = name
        self.surname = surname
        self.patronymic = patronymic
        self.sex = sex
        self.birthday = birthday
        self.data_death = data_death

    # Метод для получения полного имени по которому потом будет производиться поиск человека в базе данных
    def get_full_name(self):
        return f'{self.surname} {self.name} {self.patronymic}'.lower()

    # СтатикМетод для перевода даты в формат datetime для дальнейшего вычисления возраста человека
    @staticmethod
    def valid_date(data):
        valid_data_list = []
        date = data.split('.')
        for i in date:
            valid_data_list.append(int(i))
        date_time = datetime.date(valid_data_list[2], valid_data_list[1], valid_data_list[0])
        return date_time

    # СтатикМетод для проверки ввода на NoneType и замены его на '__' для корректной работы метода __str__
    @staticmethod
    def valid_none(word):
        if word is None:
            return "__"
        return word

    # Метод для расчета возраста
    def calculate_age(self):
        if self.data_death is None:
            today = datetime.date.today()
        else:
            today = datetime.datetime.strptime(self.data_death.replace('.', '/', 2), '%d/%m/%Y').date()
        born = datetime.datetime.strptime(self.birthday.replace('.', '/', 2), '%d/%m/%Y').date()
        return today.year - born.year - ((today.month, today.day) < (born.month, born.day))

    # Делаем изумительно красивы и продуманный вывод;) Все в одну строку, но читаемо и понятно
    def __str__(self):
        return f'| Имя: {self.name:{self.len_name}} | ' \
               f'Фамилия: {self.valid_none(self.surname):{self.len_surname}} | ' \
               f'Отчество: {self.valid_none(self.patronymic):{self.len_patronymic}} | '\
               f'Пол:{self.sex} | ' \
               f'Возраст: {self.calculate_age()} | ' \
               f'Родил{"cя: "  if self.sex in "М" else "ась:"} {self.birthday} | ' \
               f'Умер{":  " if self.sex in "М" else "ла:"} {self.valid_none(self.data_death):10} |'


class PersonList(object):
    file_name = None

    def __init__(self):
        self.persons = []

    # Метод для вывода списка людей в строку
    def __str__(self):
        return f'Список людей: {self.persons}'

    # Метод для добавления человека в список
    def add_person(self, person):
        self.persons.append(person)
        self.max_len_name()
        self.max_len_surname()
        self.max_len_patronymic()

    # Метод для сохранения всего нашего файла в формате xlsx с кастомным или уже существующим именем
    def save(self, file_name):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Люди'
        sheet.append(['Имя', 'Фамилия', 'Отчество', 'Пол', 'Дата рождения', 'Дата смерти'])
        for person in self.persons:
            sheet.append([person.name, person.surname, person.patronymic, person.sex, person.birthday,
                          person.data_death])
        wb.save(file_name)
        wb.close()
        self.persons = []
        print('-' * 30)
        print("Файл сохранился")

    # Метод для загрузки файла с уже существующим именем
    def load(self, file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb['Люди']
        for row in sheet.iter_rows(min_row=2):
            person = Person(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
            self.persons.append(person)
        wb.close()

        self.max_len_name()
        self.max_len_surname()
        self.max_len_patronymic()

        self.file_name = file_name
        print('-' * 30)
        print('Загружена база данных')

    # Методы перезаписи длинны атрибутов класса Person
    def max_len_name(self):
        max_len = 0
        for item in self.persons:
            if len(item.name) > max_len:
                max_len = len(item.name)
        Person.len_name = max_len

    def max_len_surname(self):
        max_len = 0
        for item in self.persons:
            if len(item.surname if item.surname is not None else '') > max_len:
                max_len = len(item.surname)
        Person.len_surname = max_len

    def max_len_patronymic(self):
        max_len = 0
        for item in self.persons:
            if len(item.patronymic if item.patronymic is not None else '') > max_len:
                max_len = len(item.patronymic)
        Person.len_patronymic = max_len

    # Метод для вывода информации о людях в консоль, напоменаем.. 'изумительно красиво и продуманно'
    def get_info(self):
        for person in self.persons:
            print(person)

    # Метод для поиска людей по полному имени или его части
    def find_persons(self, search_name):
        persons = []
        for person in self.persons:
            if search_name in person.get_full_name():
                persons.append(person)
        return persons

    # метод для удаления человека из списка по имени, фамилии и дате рождения
    def delete_person(self, name, surname, birthday):
        for person in self.persons:
            if person.name == name and person.surname == surname and person.birthday == birthday:
                self.persons.remove(person)
                return True


def get_valid_str(input_date: str):
    return True if input_date.isalpha() else False


def get_valid_date(input_date: str):
    return input_date.replace(" ", ".", 2).replace("/", ".", 2).replace("-", ".", 2)
